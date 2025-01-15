import streamlit as st
import requests
import pywhatkit
import time
from openpyxl import load_workbook
import os

def download_image(url, image_path):
    try:
        response = requests.get(url, stream=True, timeout=15)
        response.raise_for_status()  # Raise an exception for bad status codes (4xx or 5xx)
        with open(image_path, "wb") as file:
            for chunk in response.iter_content(1024):
                file.write(chunk)
        st.write(f"Image downloaded successfully to {image_path}!")
        return True
    except requests.exceptions.RequestException as e:
        st.error(f"A request error occurred while downloading the image from {url}: {e}")
        return False
    except Exception as e:
        st.exception(f"An unexpected error occurred while downloading the image from {url}: {e}")
        return False

def send_image_with_caption(phone_number, image_path, caption):
    try:
        st.write(f"Initiating sending image {image_path} to {phone_number}...")
        pywhatkit.sendwhats_image(phone_number, image_path, caption=caption, wait_time=30)
        st.write(f"Image {image_path} sent successfully to {phone_number}!")
        st.info("Please manually close the WhatsApp Web tab after sending.")  # User instruction
        return True
    except Exception as e:
        st.exception(f"An error occurred while sending the image to {phone_number}: {e}")
        return False

def clear_images(customer_name):
    try:
        before_path = f"before_{customer_name}.png"
        after_path = f"after_{customer_name}.png"
        if os.path.exists(before_path):
            os.remove(before_path)
            st.write(f"Removed {before_path}")
        if os.path.exists(after_path):
            os.remove(after_path)
            st.write(f"Removed {after_path}")
    except Exception as e:
        st.exception(f"Error while clearing images: {e}")

st.title("WhatsApp Image Sender")

uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx"])

if uploaded_file is not None:
    try:
        workbook = load_workbook(uploaded_file)
        sheet = workbook.active
        st.write("File uploaded successfully!")

        if st.button("Start Sending"):
            with st.spinner("Processing..."):
                for row in sheet.iter_rows(min_row=2):
                    try:
                        phone_number = row[0].value
                        customer_name = row[1].value
                        before_url = row[2].value
                        after_url = row[3].value

                        if phone_number is None or customer_name is None or before_url is None or after_url is None:
                            st.warning(f"Skipping row with missing data: {row}")
                            continue

                        if isinstance(phone_number, float):
                            phone_number = int(phone_number)

                        phone_number = str(phone_number).replace(" ", "").replace("-", "")

                        if not phone_number.startswith("+"):
                            if not phone_number.startswith("91"):
                                phone_number = "+91" + phone_number

                        st.write(f"Processing: Phone: {phone_number}, Name: {customer_name}, Before: {before_url}, After: {after_url}")

                        download_successful_before = download_image(before_url, f"before_{customer_name}.png")
                        download_successful_after = download_image(after_url, f"after_{customer_name}.png")

                        if download_successful_before and download_successful_after:
                            caption = f"Hi {customer_name}, here are your images!"
                            send_successful_before = send_image_with_caption(phone_number, f"before_{customer_name}.png", caption)
                            send_successful_after = send_image_with_caption(phone_number, f"after_{customer_name}.png", caption)

                            if not send_successful_before or not send_successful_after:
                                st.error(f"Error sending to {phone_number}. Check WhatsApp Web, number validity, and internet connection.")
                            else:
                                time.sleep(20) # Reduced sleep time after each successful send
                        else:
                            st.error(f"Download failed for {customer_name}. Check URLs and internet connection.")

                        clear_images(customer_name)
                    except Exception as e:
                        st.exception(f"An error occurred while processing a row: {e}")
                        clear_images(customer_name)
                        continue
                st.success("Finished processing the Excel file.")
    except Exception as e:
        st.exception(f"An error occurred while processing the Excel file: {e}")